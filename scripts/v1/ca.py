import datetime

import openpyxl

__all__ = [
    "ContactAngleXLS",
]


def _cell_as_str(sheet, row, col):
    # openpyxl uses 1-based row/col indexing; row and col here are 0-based
    cell = sheet.cell(row + 1, col + 1)
    value = cell.value
    if value is None:
        return ""
    elif isinstance(value, bool):
        return str(value)
    elif isinstance(value, datetime.datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        else:
            return value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    elif isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    else:
        return str(value)


class ContactAngleXLS:
    FIELD_COL, VALUE_COL = 4, 5
    NUM_FIELDS = 18
    PAD_ROWS = 3
    INIT_ROW = 7

    def __init__(self, xls_file):
        # Pass a file object to bypass openpyxl's extension check (.xls is
        # rejected by name even when the file is actually xlsx/ZIP format)
        with open(xls_file, "rb") as f:
            self.workbook = openpyxl.load_workbook(f, data_only=True)
        self.sheet = self.workbook.worksheets[0]

    def __iter__(self):
        i = self.INIT_ROW
        while i + self.NUM_FIELDS <= self.sheet.max_row:
            data = dict()
            for _ in range(self.NUM_FIELDS):
                field = _cell_as_str(self.sheet, i, self.FIELD_COL)
                value = _cell_as_str(self.sheet, i, self.VALUE_COL)
                data[field.lstrip()] = value
                i += 1
            yield data
            i += self.PAD_ROWS
